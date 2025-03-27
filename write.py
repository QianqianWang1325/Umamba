from openpyxl import load_workbook
from openpyxl.styles import Alignment


def excel_write(begin_row,begin_line,value_list):
    workbook=load_workbook("./test.xlsx")
    names= workbook.get_sheet_names()
    booksheet = workbook.get_sheet_by_name(names[0])
    for i in range(len(value_list)):
        booksheet.cell(begin_row,begin_line + i).value = value_list[i]
        booksheet.cell(begin_row,begin_line + i).alignment = Alignment(horizontal='center', vertical='center')
    workbook.save("./test.xlsx")

